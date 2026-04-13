"""
exports/excel_exporter.py
--------------------------
All Excel export logic using openpyxl.

Exports available:
  - export_patients()           — full patient register
  - export_visits()             — consultation records (date-range filterable)
  - export_inventory()          — medicines + equipment on separate sheets
  - export_expiry_report()      — expired + expiring soon medicines
  - export_weekly_report()      — combined activity summary for a week
  - export_disease_summary()    — visit counts per disease category

Usage:
    from exports.excel_exporter import export_visits
    path = export_visits(date_from="2024-01-01", date_to="2024-01-31")
    # → Returns the saved file path as a string

Each function returns the output file path so the caller (UI) can
open a file-save dialog or directly show the path to the user.
"""

import os
from datetime import date, datetime
from typing import Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

# ── DB query functions ────────────────────────────────────────────────────────
from database.patient_queries   import search_patients, get_patient_stats
from database.visit_queries     import (
    get_visits_for_export, get_disease_distribution, get_visit_stats
)
from database.inventory_queries import (
    get_medicines_for_export, get_all_equipment,
    get_expiring_soon, get_expired_medicines,
    get_inventory_stats, get_dispense_log
)

# ── Model classes (for to_export_row()) ──────────────────────────────────────
from models.patient   import Patient
from models.visit     import Visit
from models.inventory import Medicine, Equipment, DispenseRecord


# ─────────────────────────────────────────────────────────────────────────────
#  DEFAULT OUTPUT DIRECTORY
# ─────────────────────────────────────────────────────────────────────────────

DEFAULT_EXPORT_DIR = os.path.join(os.path.expanduser("~"), "Documents", "ClinicExports")


def _ensure_dir(path: str) -> str:
    os.makedirs(path, exist_ok=True)
    return path


def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")


def _output_path(filename: str, export_dir: str = None) -> str:
    directory = export_dir or DEFAULT_EXPORT_DIR
    _ensure_dir(directory)
    return os.path.join(directory, filename)


# ─────────────────────────────────────────────────────────────────────────────
#  STYLE CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

FONT_NAME = "Arial"

# Header row
HEADER_FONT       = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
HEADER_FILL_BLUE  = PatternFill("solid", fgColor="0F4C81")   # deep blue
HEADER_FILL_TEAL  = PatternFill("solid", fgColor="0D9488")   # teal — inventory
HEADER_FILL_PLUM  = PatternFill("solid", fgColor="6B21A8")   # purple — visits

# Data rows
DATA_FONT         = Font(name=FONT_NAME, size=10)
ALT_FILL          = PatternFill("solid", fgColor="F8FAFC")   # very light grey

# Status fills
FILL_EXPIRED      = PatternFill("solid", fgColor="FEE2E2")   # red tint
FILL_EXPIRING     = PatternFill("solid", fgColor="FEF3C7")   # yellow tint
FILL_LOW_STOCK    = PatternFill("solid", fgColor="FEF3C7")
FILL_OK           = PatternFill("solid", fgColor="DCFCE7")   # green tint

FONT_RED          = Font(name=FONT_NAME, color="DC2626", bold=True, size=10)
FONT_ORANGE       = Font(name=FONT_NAME, color="D97706", bold=True, size=10)
FONT_GREEN        = Font(name=FONT_NAME, color="16A34A", size=10)

# Summary / title
TITLE_FONT        = Font(name=FONT_NAME, bold=True, size=13, color="0F172A")
SUBTITLE_FONT     = Font(name=FONT_NAME, size=10, color="64748B")

CENTER_ALIGN      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN        = Alignment(horizontal="left",   vertical="center", wrap_text=True)

THIN_BORDER_SIDE  = Side(style="thin", color="E2E8F0")
THIN_BORDER       = Border(
    left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE,  bottom=THIN_BORDER_SIDE
)


# ─────────────────────────────────────────────────────────────────────────────
#  SHARED HELPERS
# ─────────────────────────────────────────────────────────────────────────────

def _write_title_block(ws, title: str, subtitle: str, col_count: int):
    """
    Writes a 2-row branded title block at the top of a worksheet.
    Caller should start data from row 4 onwards.
    """
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1,   end_column=col_count)
    ws.merge_cells(start_row=2, start_column=1,
                   end_row=2,   end_column=col_count)

    t_cell = ws.cell(row=1, column=1, value=title)
    t_cell.font      = TITLE_FONT
    t_cell.alignment = LEFT_ALIGN
    t_cell.fill      = PatternFill("solid", fgColor="E0F2FE")

    s_cell = ws.cell(row=2, column=1, value=subtitle)
    s_cell.font      = SUBTITLE_FONT
    s_cell.alignment = LEFT_ALIGN

    ws.row_dimensions[1].height = 24
    ws.row_dimensions[2].height = 16


def _write_headers(ws, headers: list[str], row: int, fill: PatternFill):
    """Write a styled header row."""
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col, value=header)
        cell.font      = HEADER_FONT
        cell.fill      = fill
        cell.alignment = CENTER_ALIGN
        cell.border    = THIN_BORDER
    ws.row_dimensions[row].height = 20


def _write_data_rows(
    ws,
    rows: list[dict],
    headers: list[str],
    start_row: int,
    row_styler=None,
):
    """
    Write data rows from a list of dicts.
    headers defines the key order and column assignment.
    row_styler(ws, excel_row_num, data_dict) → optional per-row styling hook.
    """
    for r_idx, row_data in enumerate(rows):
        excel_row = start_row + r_idx
        fill = ALT_FILL if r_idx % 2 == 0 else None

        for c_idx, key in enumerate(headers, start=1):
            val  = row_data.get(key, "")
            cell = ws.cell(row=excel_row, column=c_idx, value=val)
            cell.font      = DATA_FONT
            cell.alignment = LEFT_ALIGN
            cell.border    = THIN_BORDER
            if fill:
                cell.fill = fill

        if row_styler:
            row_styler(ws, excel_row, row_data)


def _autofit_columns(ws, min_width: int = 10, max_width: int = 50):
    """Auto-size column widths based on content."""
    for col_cells in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(min_width, min(max_len + 4, max_width))


def _freeze_header(ws, freeze_row: int = 4):
    """Freeze rows above the data so headers stay visible when scrolling."""
    ws.freeze_panes = ws.cell(row=freeze_row, column=1)


# ─────────────────────────────────────────────────────────────────────────────
#  EXPORT: PATIENTS
# ─────────────────────────────────────────────────────────────────────────────

def export_patients(
    patient_type: str = None,
    export_dir: str = None,
) -> str:
    """
    Export the patient register to Excel.
    patient_type = None → all patients
    Returns the saved file path.
    """
    patients_raw = search_patients(patient_type=patient_type, limit=100000)
    patients     = [Patient.from_dict(p) for p in patients_raw]
    rows         = [p.to_export_row() for p in patients]

    headers = list(rows[0].keys()) if rows else [
        "ID", "SAP ID", "Name", "Type", "School",
        "Age", "Gender", "Blood Group", "Mobile", "Address", "Registered On"
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Patients"

    label = patient_type or "All"
    _write_title_block(
        ws,
        f"Patient Register — {label}",
        f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  Total: {len(rows)}",
        len(headers)
    )
    _write_headers(ws, headers, row=3, fill=HEADER_FILL_BLUE)
    _write_data_rows(ws, rows, headers, start_row=4)
    _autofit_columns(ws)
    _freeze_header(ws, 4)

    ws.sheet_view.showGridLines = False

    filename = f"Patients_{label}_{_timestamp()}.xlsx"
    path = _output_path(filename, export_dir)
    wb.save(path)
    return path


# ─────────────────────────────────────────────────────────────────────────────
#  EXPORT: VISITS / CONSULTATIONS
# ─────────────────────────────────────────────────────────────────────────────

def export_visits(
    date_from: str = None,
    date_to: str = None,
    export_dir: str = None,
) -> str:
    """
    Export consultation records, optionally filtered by date range.
    date_from / date_to: 'YYYY-MM-DD'
    """
    visits_raw = get_visits_for_export(date_from=date_from, date_to=date_to)
    visits     = [Visit.from_dict(v) for v in visits_raw]
    rows       = [v.to_export_row() for v in visits]

    headers = list(rows[0].keys()) if rows else [
        "Visit ID", "Date", "Visit Type", "SAP ID", "Patient Name",
        "Patient Type", "Disease Category", "Chief Complaint", "Diagnosis",
        "Investigations", "Treatment", "Prescription", "Referral",
        "Rest Days", "Medical Leave", "Ambulance Used", "Follow-up Date", "Notes"
    ]

    wb = Workbook()
    ws = wb.active
    ws.title = "Consultations"

    date_range = ""
    if date_from or date_to:
        date_range = f"  |  {date_from or '—'} to {date_to or '—'}"

    _write_title_block(
        ws,
        f"Consultation Records{date_range}",
        f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  Total: {len(rows)}",
        len(headers)
    )
    _write_headers(ws, headers, row=3, fill=HEADER_FILL_PLUM)

    def visit_styler(ws, excel_row, row_data):
        vtype = row_data.get("Visit Type", "")
        if vtype == "Emergency":
            for col in range(1, len(headers) + 1):
                ws.cell(row=excel_row, column=col).font = Font(
                    name=FONT_NAME, size=10, color="DC2626"
                )

    _write_data_rows(ws, rows, headers, start_row=4, row_styler=visit_styler)
    _autofit_columns(ws)
    _freeze_header(ws, 4)
    ws.sheet_view.showGridLines = False

    label = f"{date_from or 'all'}_{date_to or 'all'}"
    filename = f"Consultations_{label}_{_timestamp()}.xlsx"
    path = _output_path(filename, export_dir)
    wb.save(path)
    return path


# ─────────────────────────────────────────────────────────────────────────────
#  EXPORT: INVENTORY
# ─────────────────────────────────────────────────────────────────────────────

def export_inventory(export_dir: str = None) -> str:
    """
    Full inventory export: Sheet 1 = All Medicines, Sheet 2 = Equipment.
    Expired medicines are highlighted in red, expiring soon in yellow.
    """
    medicines_raw  = get_medicines_for_export()
    equipment_raw  = get_all_equipment()
    medicines      = [Medicine.from_dict(m) for m in medicines_raw]
    equipment_list = [Equipment.from_dict(e) for e in equipment_raw]

    wb = Workbook()

    # ── Sheet 1: Medicines ──────────────────────────────────────────────────
    ws_med = wb.active
    ws_med.title = "Medicines"

    med_rows = [m.to_export_row() for m in medicines]
    med_hdrs = list(med_rows[0].keys()) if med_rows else [
        "ID", "Name", "Subtype", "Batch Number", "Supplier",
        "Stock Received", "Current Stock", "Low Stock Alert At", "Stock Status",
        "Mfg Date", "Expiry Date", "Expiry Status",
        "Days To Expiry", "Dispensed After Expiry", "Notes"
    ]

    _write_title_block(
        ws_med,
        "Medicine Inventory",
        f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  Total: {len(medicines)}",
        len(med_hdrs)
    )
    _write_headers(ws_med, med_hdrs, row=3, fill=HEADER_FILL_TEAL)

    def medicine_styler(ws, excel_row, row_data):
        status = row_data.get("Expiry Status", "")
        stock_status = row_data.get("Stock Status", "")

        if "Expired" in status:
            for col in range(1, len(med_hdrs) + 1):
                ws.cell(row=excel_row, column=col).fill = FILL_EXPIRED
            # Bold red on expiry date column
            exp_col = med_hdrs.index("Expiry Date") + 1 if "Expiry Date" in med_hdrs else None
            if exp_col:
                ws.cell(row=excel_row, column=exp_col).font = FONT_RED
        elif "Expiring" in status:
            for col in range(1, len(med_hdrs) + 1):
                ws.cell(row=excel_row, column=col).fill = FILL_EXPIRING
            exp_col = med_hdrs.index("Expiry Date") + 1 if "Expiry Date" in med_hdrs else None
            if exp_col:
                ws.cell(row=excel_row, column=exp_col).font = FONT_ORANGE

        if stock_status in ("Low Stock", "Out of Stock"):
            cur_col = med_hdrs.index("Current Stock") + 1 if "Current Stock" in med_hdrs else None
            if cur_col:
                ws.cell(row=excel_row, column=cur_col).font = FONT_RED

    _write_data_rows(ws_med, med_rows, med_hdrs, start_row=4, row_styler=medicine_styler)
    _autofit_columns(ws_med)
    _freeze_header(ws_med, 4)
    ws_med.sheet_view.showGridLines = False

    # ── Sheet 2: Equipment ───────────────────────────────────────────────────
    ws_eq = wb.create_sheet("Equipment")
    eq_rows = [e.to_export_row() for e in equipment_list]
    eq_hdrs = list(eq_rows[0].keys()) if eq_rows else [
        "ID", "Name", "Category", "Quantity",
        "Disposal Required", "Purchase Date", "Last Serviced", "Notes"
    ]

    _write_title_block(
        ws_eq,
        "Equipment & Instruments",
        f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  Total: {len(equipment_list)}",
        len(eq_hdrs)
    )
    _write_headers(ws_eq, eq_hdrs, row=3, fill=HEADER_FILL_BLUE)

    def equip_styler(ws, excel_row, row_data):
        if row_data.get("Disposal Required") == "Yes":
            dis_col = eq_hdrs.index("Disposal Required") + 1
            ws.cell(row=excel_row, column=dis_col).font = FONT_ORANGE

    _write_data_rows(ws_eq, eq_rows, eq_hdrs, start_row=4, row_styler=equip_styler)
    _autofit_columns(ws_eq)
    _freeze_header(ws_eq, 4)
    ws_eq.sheet_view.showGridLines = False

    filename = f"Inventory_{_timestamp()}.xlsx"
    path = _output_path(filename, export_dir)
    wb.save(path)
    return path


# ─────────────────────────────────────────────────────────────────────────────
#  EXPORT: EXPIRY REPORT
# ─────────────────────────────────────────────────────────────────────────────

def export_expiry_report(export_dir: str = None) -> str:
    """
    Two-sheet report:
      Sheet 1 → Expired medicines
      Sheet 2 → Expiring within 2 months
    """
    expired_raw  = get_expired_medicines()
    expiring_raw = get_expiring_soon(months=2)
    expired      = [Medicine.from_dict(m) for m in expired_raw]
    expiring     = [Medicine.from_dict(m) for m in expiring_raw]

    wb = Workbook()

    med_hdrs = [
        "ID", "Name", "Subtype", "Batch Number", "Supplier",
        "Current Stock", "Mfg Date", "Expiry Date",
        "Dispensed After Expiry", "Notes"
    ]

    for sheet_data, sheet_title, fill, font in [
        (expired,  "Expired Medicines",        HEADER_FILL_PLUM, FONT_RED),
        (expiring, "Expiring Soon (2 months)", HEADER_FILL_TEAL, FONT_ORANGE),
    ]:
        ws = wb.active if sheet_title == "Expired Medicines" else wb.create_sheet()
        ws.title = sheet_title[:31]

        rows = [m.to_export_row() for m in sheet_data]
        # Filter to relevant columns
        filtered_rows = [
            {k: r.get(k, "") for k in med_hdrs} for r in rows
        ]

        _write_title_block(
            ws,
            sheet_title,
            f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  Count: {len(rows)}",
            len(med_hdrs)
        )
        _write_headers(ws, med_hdrs, row=3, fill=fill)

        if sheet_title.startswith("Expired"):
            row_fill = FILL_EXPIRED
        else:
            row_fill = FILL_EXPIRING

        def expiry_styler(ws, excel_row, row_data, _font=font, _fill=row_fill):
            for col in range(1, len(med_hdrs) + 1):
                ws.cell(row=excel_row, column=col).fill = _fill

        _write_data_rows(ws, filtered_rows, med_hdrs, start_row=4,
                         row_styler=expiry_styler)
        _autofit_columns(ws)
        _freeze_header(ws, 4)
        ws.sheet_view.showGridLines = False

    filename = f"ExpiryReport_{_timestamp()}.xlsx"
    path = _output_path(filename, export_dir)
    wb.save(path)
    return path


# ─────────────────────────────────────────────────────────────────────────────
#  EXPORT: DISEASE SUMMARY
# ─────────────────────────────────────────────────────────────────────────────

def export_disease_summary(
    date_from: str = None,
    date_to: str = None,
    export_dir: str = None,
) -> str:
    """
    One-sheet report: visits grouped and counted per disease category.
    """
    distribution = get_disease_distribution(date_from=date_from, date_to=date_to)
    total_visits  = sum(r.get("visit_count", 0) for r in distribution)

    # Add percentage column
    rows = []
    for r in sorted(distribution, key=lambda x: x.get("visit_count", 0), reverse=True):
        count = r.get("visit_count", 0)
        pct   = f"{(count / total_visits * 100):.1f}%" if total_visits else "0%"
        rows.append({
            "Disease Category": r.get("category_name", "—"),
            "Visit Count":      count,
            "Percentage":       pct,
        })

    headers = ["Disease Category", "Visit Count", "Percentage"]

    wb  = Workbook()
    ws  = wb.active
    ws.title = "Disease Summary"

    date_range = f"  |  {date_from or 'All time'} to {date_to or 'today'}"
    _write_title_block(
        ws,
        f"Disease Distribution{date_range}",
        f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  Total Visits: {total_visits}",
        len(headers)
    )
    _write_headers(ws, headers, row=3, fill=HEADER_FILL_TEAL)
    _write_data_rows(ws, rows, headers, start_row=4)
    _autofit_columns(ws)
    _freeze_header(ws, 4)
    ws.sheet_view.showGridLines = False

    # Bold + green the top disease
    if rows:
        for col in range(1, len(headers) + 1):
            ws.cell(row=4, column=col).font = Font(
                name=FONT_NAME, bold=True, size=10, color="0F4C81"
            )

    filename = f"DiseaseSummary_{_timestamp()}.xlsx"
    path = _output_path(filename, export_dir)
    wb.save(path)
    return path


# ─────────────────────────────────────────────────────────────────────────────
#  EXPORT: WEEKLY REPORT
# ─────────────────────────────────────────────────────────────────────────────

def export_weekly_report(
    week_start: str = None,   # 'YYYY-MM-DD' — defaults to Monday of current week
    export_dir: str = None,
) -> str:
    """
    5-sheet comprehensive weekly report:
      1. Summary          — stats overview
      2. Consultations    — all visits that week
      3. Disease Breakdown— category distribution
      4. Medicines        — current inventory snapshot
      5. Dispense Log     — what was dispensed that week
    """
    from datetime import timedelta

    today  = date.today()
    if week_start:
        w_start = date.fromisoformat(week_start)
    else:
        w_start = today - timedelta(days=today.weekday())  # Monday
    w_end = w_start + timedelta(days=6)

    w_start_str = w_start.isoformat()
    w_end_str   = w_end.isoformat()
    week_label  = f"{w_start.strftime('%d %b')} – {w_end.strftime('%d %b %Y')}"

    wb = Workbook()

    # ── Sheet 1: Summary ─────────────────────────────────────────────────────
    ws_sum = wb.active
    ws_sum.title = "Summary"

    visit_stats = get_visit_stats(date_from=w_start_str, date_to=w_end_str)
    inv_stats   = get_inventory_stats()
    pat_stats   = get_patient_stats()

    ws_sum.merge_cells("A1:D1")
    ws_sum["A1"] = f"Weekly Clinic Report  —  {week_label}"
    ws_sum["A1"].font      = Font(name=FONT_NAME, bold=True, size=14, color="0F172A")
    ws_sum["A1"].fill      = PatternFill("solid", fgColor="E0F2FE")
    ws_sum["A1"].alignment = LEFT_ALIGN
    ws_sum.row_dimensions[1].height = 28

    ws_sum.merge_cells("A2:D2")
    ws_sum["A2"] = f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}"
    ws_sum["A2"].font = SUBTITLE_FONT

    summary_rows = [
        ("VISIT STATISTICS",       ""),
        ("Total Visits This Week",   visit_stats.get("total_visits", 0)),
        ("Walk-ins",                 visit_stats.get("walk_in", 0)),
        ("Scheduled",                visit_stats.get("scheduled", 0)),
        ("Emergency",                visit_stats.get("emergency", 0)),
        ("With Referral",            visit_stats.get("with_referral", 0)),
        ("Medical Leave Issued",     visit_stats.get("with_medical_leave", 0)),
        ("Ambulance Used",           visit_stats.get("with_ambulance", 0)),
        ("",                         ""),
        ("PATIENT REGISTER",         ""),
        ("Total Patients",           pat_stats.get("total", 0)),
        ("Students",                 pat_stats.get("students", 0)),
        ("Staff",                    pat_stats.get("staff", 0)),
        ("",                         ""),
        ("INVENTORY STATUS",         ""),
        ("Total Medicines",          inv_stats.get("total_medicines", 0)),
        ("Expired Medicines",        inv_stats.get("expired_count", 0)),
        ("Expiring Within 2 Months", inv_stats.get("expiring_soon_count", 0)),
        ("Low Stock Items",          inv_stats.get("low_stock_count", 0)),
        ("Total Equipment",          inv_stats.get("total_equipment", 0)),
        ("Equipment Needing Disposal",inv_stats.get("disposal_needed_count", 0)),
    ]

    for r_idx, (label, value) in enumerate(summary_rows, start=4):
        ws_sum.cell(row=r_idx, column=1, value=label).font = (
            Font(name=FONT_NAME, bold=True, size=10, color="0F4C81")
            if str(value) == "" and label
            else DATA_FONT
        )
        cell_val = ws_sum.cell(row=r_idx, column=2, value=value)
        cell_val.font = DATA_FONT
        if isinstance(value, int) and value > 0 and "Expired" in label:
            cell_val.font = FONT_RED
        ws_sum.row_dimensions[r_idx].height = 18

    ws_sum.column_dimensions["A"].width = 36
    ws_sum.column_dimensions["B"].width = 14
    ws_sum.sheet_view.showGridLines = False

    # ── Sheet 2: Consultations ────────────────────────────────────────────────
    ws_vis = wb.create_sheet("Consultations")
    visits_raw = get_visits_for_export(date_from=w_start_str, date_to=w_end_str)
    visits     = [Visit.from_dict(v) for v in visits_raw]
    vis_rows   = [v.to_export_row() for v in visits]

    vis_hdrs = list(vis_rows[0].keys()) if vis_rows else [
        "Visit ID", "Date", "Visit Type", "SAP ID", "Patient Name",
        "Disease Category", "Diagnosis", "Treatment",
        "Medical Leave", "Ambulance Used"
    ]

    _write_title_block(
        ws_vis,
        f"Consultations  —  {week_label}",
        f"Total: {len(vis_rows)}",
        len(vis_hdrs)
    )
    _write_headers(ws_vis, vis_hdrs, row=3, fill=HEADER_FILL_PLUM)
    _write_data_rows(ws_vis, vis_rows, vis_hdrs, start_row=4)
    _autofit_columns(ws_vis)
    _freeze_header(ws_vis, 4)
    ws_vis.sheet_view.showGridLines = False

    # ── Sheet 3: Disease Breakdown ────────────────────────────────────────────
    ws_dis = wb.create_sheet("Disease Breakdown")
    dist_raw = get_disease_distribution(date_from=w_start_str, date_to=w_end_str)
    total    = sum(r.get("visit_count", 0) for r in dist_raw)
    dis_rows = [
        {
            "Disease Category": r.get("category_name", "—"),
            "Visit Count":      r.get("visit_count", 0),
            "Percentage":       f"{r.get('visit_count',0)/total*100:.1f}%" if total else "0%",
        }
        for r in sorted(dist_raw, key=lambda x: x.get("visit_count", 0), reverse=True)
    ]
    dis_hdrs = ["Disease Category", "Visit Count", "Percentage"]

    _write_title_block(
        ws_dis, f"Disease Distribution  —  {week_label}",
        f"Total Visits: {total}", len(dis_hdrs)
    )
    _write_headers(ws_dis, dis_hdrs, row=3, fill=HEADER_FILL_TEAL)
    _write_data_rows(ws_dis, dis_rows, dis_hdrs, start_row=4)
    _autofit_columns(ws_dis)
    _freeze_header(ws_dis, 4)
    ws_dis.sheet_view.showGridLines = False

    # ── Sheet 4: Medicine Inventory Snapshot ──────────────────────────────────
    ws_inv = wb.create_sheet("Medicines")
    med_raw  = get_medicines_for_export()
    meds     = [Medicine.from_dict(m) for m in med_raw]
    med_rows = [m.to_export_row() for m in meds]
    med_hdrs = list(med_rows[0].keys()) if med_rows else []

    _write_title_block(
        ws_inv, "Medicine Inventory Snapshot",
        f"As of: {datetime.now().strftime('%d %b %Y %H:%M')}  |  Total: {len(meds)}",
        len(med_hdrs)
    )
    _write_headers(ws_inv, med_hdrs, row=3, fill=HEADER_FILL_TEAL)
    _write_data_rows(ws_inv, med_rows, med_hdrs, start_row=4)
    _autofit_columns(ws_inv)
    _freeze_header(ws_inv, 4)
    ws_inv.sheet_view.showGridLines = False

    # ── Sheet 5: Dispense Log ─────────────────────────────────────────────────
    ws_disp = wb.create_sheet("Dispense Log")
    disp_raw  = get_dispense_log(date_from=w_start_str, date_to=w_end_str)
    dispenses = [DispenseRecord.from_dict(d) for d in disp_raw]
    disp_rows = [d.to_export_row() for d in dispenses]
    disp_hdrs = list(disp_rows[0].keys()) if disp_rows else [
        "Dispense ID", "Medicine", "Quantity",
        "Patient", "SAP ID", "Dispensed By", "Dispensed At", "Notes"
    ]

    _write_title_block(
        ws_disp, f"Dispense Log  —  {week_label}",
        f"Total Dispenses: {len(dispenses)}",
        len(disp_hdrs)
    )
    _write_headers(ws_disp, disp_hdrs, row=3, fill=HEADER_FILL_BLUE)
    _write_data_rows(ws_disp, disp_rows, disp_hdrs, start_row=4)
    _autofit_columns(ws_disp)
    _freeze_header(ws_disp, 4)
    ws_disp.sheet_view.showGridLines = False

    filename = f"WeeklyReport_{w_start_str}_{_timestamp()}.xlsx"
    path = _output_path(filename, export_dir)
    wb.save(path)
    return path