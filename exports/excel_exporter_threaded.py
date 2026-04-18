"""
exports/excel_exporter.py
--------------------------
All Excel export logic using openpyxl. Moved to QThread to prevent UI blocking.
Streams data dynamically in chunks.
"""

import os
from datetime import date, datetime
from typing import Optional

import openpyxl
from openpyxl import Workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side, numbers
)
from openpyxl.utils import get_column_letter

from PySide6.QtCore import QThread, Signal

# ── DB query functions ────────────────────────────────────────────────────────
from database.db_manager      import get_connection

# ─────────────────────────────────────────────────────────────────────────────
#  DEFAULT OUTPUT DIRECTORY
# ─────────────────────────────────────────────────────────────────────────────

DEFAULT_EXPORT_DIR = os.path.join(os.path.expanduser("~"), "Documents", "ClinicExports")

def _ensure_dir(path: str) -> str:
    os.makedirs(path, exist_ok=True)
    return path

def _timestamp() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")

def _output_path(filename: str, export_dir: str = None) -> str:
    directory = export_dir or DEFAULT_EXPORT_DIR
    _ensure_dir(directory)
    return os.path.join(directory, filename)

# ─────────────────────────────────────────────────────────────────────────────
#  STYLE CONSTANTS
# ─────────────────────────────────────────────────────────────────────────────

FONT_NAME = "Arial"
HEADER_FONT       = Font(name=FONT_NAME, bold=True, color="FFFFFF", size=10)
HEADER_FILL_BLUE  = PatternFill("solid", fgColor="0F4C81")
DATA_FONT         = Font(name=FONT_NAME, size=10)
ALT_FILL          = PatternFill("solid", fgColor="F8FAFC")
TITLE_FONT        = Font(name=FONT_NAME, bold=True, size=13, color="0F172A")
SUBTITLE_FONT     = Font(name=FONT_NAME, size=10, color="64748B")
CENTER_ALIGN      = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT_ALIGN        = Alignment(horizontal="left",   vertical="center", wrap_text=True)
THIN_BORDER_SIDE  = Side(style="thin", color="E2E8F0")
THIN_BORDER       = Border(
    left=THIN_BORDER_SIDE, right=THIN_BORDER_SIDE,
    top=THIN_BORDER_SIDE,  bottom=THIN_BORDER_SIDE
)

class BaseExportThread(QThread):
    progress = Signal(int, str)  # percentage, message
    finished = Signal(str)       # file_path
    error = Signal(str)          # error message

    def __init__(self, export_dir=None):
        super().__init__()
        self.export_dir = export_dir
        self.chunk_size = 1000

    def _write_title_block(self, ws, title: str, subtitle: str, col_count: int):
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_count)
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_count)
        t_cell = ws.cell(row=1, column=1, value=title)
        t_cell.font, t_cell.alignment, t_cell.fill = TITLE_FONT, LEFT_ALIGN, PatternFill("solid", fgColor="E0F2FE")
        s_cell = ws.cell(row=2, column=1, value=subtitle)
        s_cell.font, s_cell.alignment = SUBTITLE_FONT, LEFT_ALIGN
        ws.row_dimensions[1].height, ws.row_dimensions[2].height = 24, 16

    def _write_headers(self, ws, headers, row, fill):
        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font, cell.fill, cell.alignment, cell.border = HEADER_FONT, fill, CENTER_ALIGN, THIN_BORDER
        ws.row_dimensions[row].height = 20

    def _autofit_columns(self, ws):
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max(10, min(max_len + 4, 50))


class ExportPatientsThread(BaseExportThread):
    def __init__(self, patient_type=None, export_dir=None):
        super().__init__(export_dir)
        self.patient_type = patient_type

    def run(self):
        conn = None
        try:
            self.progress.emit(0, "Initializing database...")
            conn = get_connection()
            
            # Count total
            count_query = "SELECT COUNT(*) FROM patients"
            params = []
            if self.patient_type:
                count_query += " WHERE type = ?"
                params.append(self.patient_type)
                
            total = conn.execute(count_query, params).fetchone()[0]
            if total == 0:
                self.error.emit("No patients found to export.")
                return

            self.progress.emit(5, f"Found {total} patients. Creating workbook...")
            wb = Workbook()
            ws = wb.active
            ws.title = "Patients"

            headers = ["ID", "SAP ID", "Name", "Type", "School", "Age", "Gender", "Blood Group", "Mobile", "Address", "Registered On"]
            
            label = self.patient_type or "All"
            self._write_title_block(ws, f"Patient Register — {label}", f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  Total: {total}", len(headers))
            self._write_headers(ws, headers, row=3, fill=HEADER_FILL_BLUE)
            ws.freeze_panes = ws.cell(row=4, column=1)

            # Query data in chunks
            query = "SELECT id, sap_id, name, type, school, age, gender, blood_group, mobile, address, created_at FROM patients"
            if self.patient_type:
                query += " WHERE type = ?"
                
            cursor = conn.cursor()
            cursor.execute(query, params)
            
            current_row = 4
            processed = 0
            
            while True:
                chunk = cursor.fetchmany(self.chunk_size)
                if not chunk:
                    break
                    
                for row in chunk:
                    for col, val in enumerate(row, start=1):
                        cell = ws.cell(row=current_row, column=col, value=val)
                        cell.font, cell.alignment, cell.border = DATA_FONT, LEFT_ALIGN, THIN_BORDER
                        if (current_row - 4) % 2 == 0:
                            cell.fill = ALT_FILL
                    current_row += 1
                    processed += 1
                    
                pct = int(5 + (processed / total) * 85)
                self.progress.emit(pct, f"Exported {processed}/{total} patients...")

            self.progress.emit(90, "Formatting columns...")
            self._autofit_columns(ws)
            ws.sheet_view.showGridLines = False

            self.progress.emit(95, "Saving file...")
            filename = f"Patients_{label}_{_timestamp()}.xlsx"
            path = _output_path(filename, self.export_dir)
            wb.save(path)
            
            self.progress.emit(100, "Done!")
            self.finished.emit(path)
            
        except Exception as e:
            self.error.emit(str(e))
        finally:
            if conn:
                conn.close()

class ExportVisitsThread(BaseExportThread):
    def run(self):
        conn = None
        try:
            self.progress.emit(0, "Initializing database...")
            conn = get_connection()
            total = conn.execute("SELECT COUNT(*) FROM visits").fetchone()[0]
            if total == 0:
                self.error.emit("No visits found to export.")
                return

            self.progress.emit(5, f"Found {total} visits. Creating workbook...")
            wb = Workbook()
            ws = wb.active
            ws.title = "Visits"

            headers = [
                "Visit ID", "Date", "Type", "Patient SAP ID", "Patient Name",
                "Patient Type", "School", "Age", "Gender", "Category",
                "Complaint", "Diagnosis", "Investigations", "Treatment",
                "Prescription", "Referral", "Rest Days", "Medical Leave",
                "Ambulance Used", "Follow-up", "Notes"
            ]
            
            self._write_title_block(ws, "Consultation Visits", f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}  |  Total: {total}", len(headers))
            self._write_headers(ws, headers, row=3, fill=PatternFill("solid", fgColor="0D9488"))
            ws.freeze_panes = ws.cell(row=4, column=1)

            from database.visit_queries import get_visits_for_export
            visits = get_visits_for_export()
            
            current_row = 4
            processed = 0
            
            for v in visits:
                cells = [
                    v.get("id"), v.get("visit_date"), v.get("visit_type"),
                    v.get("patient_sap_id"), v.get("patient_name"), v.get("patient_type"),
                    v.get("school"), v.get("age"), v.get("gender"), v.get("disease_category"),
                    v.get("chief_complaint"), v.get("diagnosis"), v.get("investigations"),
                    v.get("treatment"), v.get("prescription"), v.get("referral"),
                    v.get("rest_days"), 1 if v.get("medical_leave") else 0,
                    1 if v.get("ambulance_used") else 0, v.get("follow_up_date"), v.get("notes")
                ]
                for col, val in enumerate(cells, start=1):
                    cell = ws.cell(row=current_row, column=col, value=val)
                    cell.font, cell.alignment, cell.border = DATA_FONT, LEFT_ALIGN, THIN_BORDER
                    if (current_row - 4) % 2 == 0:
                        cell.fill = ALT_FILL
                current_row += 1
                processed += 1
                if processed % 100 == 0:
                    pct = int(5 + (processed / total) * 85)
                    self.progress.emit(pct, f"Exported {processed}/{total} visits...")

            self.progress.emit(90, "Formatting columns...")
            self._autofit_columns(ws)
            ws.sheet_view.showGridLines = False

            self.progress.emit(95, "Saving file...")
            filename = f"Visits_{_timestamp()}.xlsx"
            path = _output_path(filename, self.export_dir)
            wb.save(path)
            
            self.progress.emit(100, "Done!")
            self.finished.emit(path)
            
        except Exception as e:
            self.error.emit(str(e))
        finally:
            if conn:
                conn.close()

class ExportInventoryThread(BaseExportThread):
    def run(self):
        try:
            self.progress.emit(0, "Fetching inventory data...")
            wb = Workbook()
            ws = wb.active
            ws.title = "Medicines"

            headers = [
                "ID", "Name", "Subtype", "Batch Number", "Stock Received",
                "Current Stock", "Min Stock Alert", "Mfg Date", "Expiry Date",
                "Dispensed After Expiry", "Supplier", "Notes", "Created At"
            ]
            
            self._write_title_block(ws, "Inventory - Medicines", f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", len(headers))
            self._write_headers(ws, headers, row=3, fill=PatternFill("solid", fgColor="D97706"))
            ws.freeze_panes = ws.cell(row=4, column=1)

            from database.inventory_queries import get_medicines_for_export, get_all_equipment
            medicines = get_medicines_for_export()
            total = len(medicines)
            if total == 0:
                self.error.emit("No medicines found to export.")
                return

            current_row = 4
            processed = 0
            
            for m in medicines:
                cells = [
                    m.get("id"), m.get("name"), m.get("subtype"), m.get("batch_number"),
                    m.get("stock_received"), m.get("current_stock"), m.get("minimum_stock_alert"),
                    m.get("mfg_date"), m.get("expiry_date"), m.get("dispensed_after_expiry"),
                    m.get("supplier"), m.get("notes"), m.get("created_at")
                ]
                for col, val in enumerate(cells, start=1):
                    cell = ws.cell(row=current_row, column=col, value=val)
                    cell.font, cell.alignment, cell.border = DATA_FONT, LEFT_ALIGN, THIN_BORDER
                    if (current_row - 4) % 2 == 0:
                        cell.fill = ALT_FILL
                current_row += 1
                processed += 1
                if processed % 50 == 0:
                    pct = int(5 + (processed / total) * 40)
                    self.progress.emit(pct, f"Exported {processed}/{total} medicines...")

            self._autofit_columns(ws)
            ws.sheet_view.showGridLines = False

            # Add Equipment Sheet
            self.progress.emit(50, "Exporting equipment...")
            ws_eq = wb.create_sheet(title="Equipment")
            headers_eq = ["ID", "Name", "Category", "Quantity", "Disposal Required", "Purchase Date", "Last Serviced", "Notes"]
            self._write_title_block(ws_eq, "Inventory - Equipment", f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", len(headers_eq))
            self._write_headers(ws_eq, headers_eq, row=3, fill=PatternFill("solid", fgColor="D97706"))
            
            equipment = get_all_equipment()
            current_row = 4
            for eq in equipment:
                cells = [
                    eq.get("id"), eq.get("name"), eq.get("category"), eq.get("quantity"),
                    1 if eq.get("disposal_required") else 0, eq.get("purchase_date"),
                    eq.get("last_serviced_date"), eq.get("notes")
                ]
                for col, val in enumerate(cells, start=1):
                    cell = ws_eq.cell(row=current_row, column=col, value=val)
                    cell.font, cell.alignment, cell.border = DATA_FONT, LEFT_ALIGN, THIN_BORDER
                    if (current_row - 4) % 2 == 0:
                        cell.fill = ALT_FILL
                current_row += 1

            self._autofit_columns(ws_eq)
            ws_eq.sheet_view.showGridLines = False

            self.progress.emit(95, "Saving file...")
            filename = f"Inventory_{_timestamp()}.xlsx"
            path = _output_path(filename, self.export_dir)
            wb.save(path)
            
            self.progress.emit(100, "Done!")
            self.finished.emit(path)
            
        except Exception as e:
            self.error.emit(str(e))

class ExportAllDataThread(BaseExportThread):
    def run(self):
        try:
            self.progress.emit(0, "Exporting all data... This might take a while.")
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet

            # 1. Patients
            self.progress.emit(10, "Exporting Patients...")
            ws_pat = wb.create_sheet(title="Patients")
            headers_pat = ["ID", "SAP ID", "Name", "Type", "School", "Age", "Gender", "Blood Group", "Mobile", "Address", "Registered On"]
            self._write_title_block(ws_pat, "All Patients", f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", len(headers_pat))
            self._write_headers(ws_pat, headers_pat, row=3, fill=HEADER_FILL_BLUE)
            ws_pat.freeze_panes = ws_pat.cell(row=4, column=1)

            conn = get_connection()
            cursor = conn.cursor()
            cursor.execute("SELECT id, sap_id, name, type, school, age, gender, blood_group, mobile, address, created_at FROM patients")
            current_row = 4
            while True:
                chunk = cursor.fetchmany(self.chunk_size)
                if not chunk: break
                for row in chunk:
                    for col, val in enumerate(row, start=1):
                        cell = ws_pat.cell(row=current_row, column=col, value=val)
                        cell.font, cell.alignment, cell.border = DATA_FONT, LEFT_ALIGN, THIN_BORDER
                        if (current_row - 4) % 2 == 0: cell.fill = ALT_FILL
                    current_row += 1
            self._autofit_columns(ws_pat)
            ws_pat.sheet_view.showGridLines = False

            # 2. Visits
            self.progress.emit(40, "Exporting Visits...")
            ws_vis = wb.create_sheet(title="Visits")
            headers_vis = [
                "Visit ID", "Date", "Type", "Patient SAP ID", "Patient Name",
                "Patient Type", "School", "Age", "Gender", "Category",
                "Complaint", "Diagnosis", "Investigations", "Treatment",
                "Prescription", "Referral", "Rest Days", "Medical Leave",
                "Ambulance Used", "Follow-up", "Notes"
            ]
            self._write_title_block(ws_vis, "All Visits", f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", len(headers_vis))
            self._write_headers(ws_vis, headers_vis, row=3, fill=PatternFill("solid", fgColor="0D9488"))
            ws_vis.freeze_panes = ws_vis.cell(row=4, column=1)

            from database.visit_queries import get_visits_for_export
            visits = get_visits_for_export()
            current_row = 4
            for v in visits:
                cells = [
                    v.get("id"), v.get("visit_date"), v.get("visit_type"),
                    v.get("patient_sap_id"), v.get("patient_name"), v.get("patient_type"),
                    v.get("school"), v.get("age"), v.get("gender"), v.get("disease_category"),
                    v.get("chief_complaint"), v.get("diagnosis"), v.get("investigations"),
                    v.get("treatment"), v.get("prescription"), v.get("referral"),
                    v.get("rest_days"), 1 if v.get("medical_leave") else 0,
                    1 if v.get("ambulance_used") else 0, v.get("follow_up_date"), v.get("notes")
                ]
                for col, val in enumerate(cells, start=1):
                    cell = ws_vis.cell(row=current_row, column=col, value=val)
                    cell.font, cell.alignment, cell.border = DATA_FONT, LEFT_ALIGN, THIN_BORDER
                    if (current_row - 4) % 2 == 0: cell.fill = ALT_FILL
                current_row += 1
            self._autofit_columns(ws_vis)
            ws_vis.sheet_view.showGridLines = False

            # 3. Medicines
            self.progress.emit(70, "Exporting Inventory...")
            ws_med = wb.create_sheet(title="Medicines")
            headers_med = [
                "ID", "Name", "Subtype", "Batch Number", "Stock Received",
                "Current Stock", "Min Stock Alert", "Mfg Date", "Expiry Date",
                "Dispensed After Expiry", "Supplier", "Notes", "Created At"
            ]
            self._write_title_block(ws_med, "Inventory - Medicines", f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", len(headers_med))
            self._write_headers(ws_med, headers_med, row=3, fill=PatternFill("solid", fgColor="D97706"))
            ws_med.freeze_panes = ws_med.cell(row=4, column=1)

            from database.inventory_queries import get_medicines_for_export, get_all_equipment
            medicines = get_medicines_for_export()
            current_row = 4
            for m in medicines:
                cells = [
                    m.get("id"), m.get("name"), m.get("subtype"), m.get("batch_number"),
                    m.get("stock_received"), m.get("current_stock"), m.get("minimum_stock_alert"),
                    m.get("mfg_date"), m.get("expiry_date"), m.get("dispensed_after_expiry"),
                    m.get("supplier"), m.get("notes"), m.get("created_at")
                ]
                for col, val in enumerate(cells, start=1):
                    cell = ws_med.cell(row=current_row, column=col, value=val)
                    cell.font, cell.alignment, cell.border = DATA_FONT, LEFT_ALIGN, THIN_BORDER
                    if (current_row - 4) % 2 == 0: cell.fill = ALT_FILL
                current_row += 1
            self._autofit_columns(ws_med)
            ws_med.sheet_view.showGridLines = False

            # 4. Equipment
            ws_eq = wb.create_sheet(title="Equipment")
            headers_eq = ["ID", "Name", "Category", "Quantity", "Disposal Required", "Purchase Date", "Last Serviced", "Notes"]
            self._write_title_block(ws_eq, "Inventory - Equipment", f"Generated: {datetime.now().strftime('%d %b %Y %H:%M')}", len(headers_eq))
            self._write_headers(ws_eq, headers_eq, row=3, fill=PatternFill("solid", fgColor="D97706"))
            ws_eq.freeze_panes = ws_eq.cell(row=4, column=1)

            equipment = get_all_equipment()
            current_row = 4
            for eq in equipment:
                cells = [
                    eq.get("id"), eq.get("name"), eq.get("category"), eq.get("quantity"),
                    1 if eq.get("disposal_required") else 0, eq.get("purchase_date"),
                    eq.get("last_serviced_date"), eq.get("notes")
                ]
                for col, val in enumerate(cells, start=1):
                    cell = ws_eq.cell(row=current_row, column=col, value=val)
                    cell.font, cell.alignment, cell.border = DATA_FONT, LEFT_ALIGN, THIN_BORDER
                    if (current_row - 4) % 2 == 0: cell.fill = ALT_FILL
                current_row += 1
            self._autofit_columns(ws_eq)
            ws_eq.sheet_view.showGridLines = False

            self.progress.emit(95, "Saving file...")
            filename = f"Clinic_Full_Export_{_timestamp()}.xlsx"
            path = _output_path(filename, self.export_dir)
            wb.save(path)
            
            self.progress.emit(100, "Done!")
            self.finished.emit(path)
            
        except Exception as e:
            self.error.emit(str(e))
        finally:
            if 'conn' in locals() and conn:
                conn.close()
